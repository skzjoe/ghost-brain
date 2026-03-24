#!/usr/bin/env python3
import argparse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl_excel_builder import new_workbook, add_sheet_from_rows, save_workbook

GREEN='385723'
BLUE='1F4E78'
ORANGE='C55A11'
PURPLE='5B2C6F'


def executive_summary_rows(project_name, customer_name, duration_weeks):
    return [
        ['Field','Value'],
        ['Project', project_name],
        ['Customer', customer_name],
        ['Project Type', 'ERPNext implementation - medium scope'],
        ['Planning Basis', 'PM-led phased delivery with scope control, stage gates, and controlled go-live'],
        ['Recommended Duration', f'{duration_weeks} weeks'],
        ['Target Scope', 'Sales, Purchase, Inventory, Accounting, Manufacturing'],
        ['Delivery Strategy', 'Phased design-build-test-cutover with explicit sign-offs'],
        ['Success Criteria', 'Scope sign-off, clean migration, UAT pass, stable go-live, hypercare closure'],
        ['Key Constraints', 'Customer decision speed, data quality, customization load, user availability'],
    ]


def scope_rows():
    return [
        ['Category','Included in Phase 1','Deferred / Out of Scope Notes'],
        ['Core modules','Sales, Purchase, Inventory, Accounting, Manufacturing','Advanced BI or non-critical custom apps go phase 2'],
        ['Master data','Customer, Supplier, Item, UOM, Warehouse, BOM, Routing, COA','Historical archive migration only if justified'],
        ['Transactions','Opening stock, AR/AP open items, open SO/PO/WO if required','Legacy detailed historical transactional migration optional'],
        ['Customizations','Only must-have fields, reports, workflow logic, and integrations','Nice-to-have enhancements after stabilization'],
        ['Training','Key user + end user training','Advanced admin playbooks can follow in phase 2'],
        ['Go-live support','Cutover + hypercare','Long-term support handled under BAU/support agreement'],
    ]


def milestone_rows():
    return [
        ['Milestone ID','Milestone','Week','Owner','Acceptance Gate','Status'],
        ['M1','Project Kickoff Completed',1,'Implementer PM + Customer PM','Project charter approved; owners assigned','Planned'],
        ['M2','Requirements & Fit-Gap Signed Off',4,'Customer PM + Key Users','Requirements, process map, fit-gap approved','Planned'],
        ['M3','Solution Blueprint Approved',5,'Sponsor + Solution Architect','Workflows, roles, integrations, migration strategy approved','Planned'],
        ['M4','Core Configuration Completed',8,'Functional Lead','Core modules configured and internally reviewed','Planned'],
        ['M5','Build Complete',10,'Technical Lead','Approved custom build deployed to UAT','Planned'],
        ['M6','Migration Dry Run Completed',10,'Data Owner + Data Consultant','Balances and key masters validated','Planned'],
        ['M7','SIT Passed',11,'Implementer Team','Critical end-to-end flows pass','Planned'],
        ['M8','UAT Signed Off',12,'Customer PM + Key Users','Business scenarios approved by named users','Planned'],
        ['M9','Go-Live Readiness Approved',13,'Sponsor + PMs','Cutover checklist and rollback approved','Planned'],
        ['M10','Production Go-Live',13,'Project Team','First live transactions executed successfully','Planned'],
        ['M11','Hypercare Closed',16,'Support Lead + Customer Owner','P1/P2 issues closed or accepted','Planned'],
    ]


def phase_plan_rows():
    return [
        ['Phase','Week Start','Week End','Primary Objective','Key Deliverables','Key Dependencies','Exit Criteria'],
        ['1. Initiation',1,1,'Stand up project governance and ownership','Charter, governance, stakeholder register, RAID log','Sponsor + PM assignment','Kickoff complete'],
        ['2. Requirements & Fit-Gap',2,4,'Define scope, process, and standard-vs-custom decisions','Requirements pack, fit-gap, process maps, master data design','Key user availability','Requirements sign-off'],
        ['3. Solution Design',4,5,'Translate requirements into solution and delivery plan','Blueprint, role matrix, migration approach, test/cutover strategy','Fit-gap decisions complete','Blueprint approved'],
        ['4. Environment Setup',5,6,'Prepare delivery environments and technical baseline','DEV/UAT/PROD ready, backups, SSL, email baseline','Infra readiness','Environment ready'],
        ['5. Configuration',6,8,'Configure ERPNext standard behavior','Configured core modules, numbering, permissions, print baseline','Blueprint approved','Core config complete'],
        ['6. Custom Build',6,10,'Deliver approved custom logic and outputs','Custom fields, scripts, reports, integrations','Requirements frozen for must-haves','Build complete'],
        ['7. Data Migration',5,10,'Prepare, cleanse, map, and validate go-live data','Templates, cleansed data, dry-run validation','Master data design','Dry run complete'],
        ['8. Testing & Training',10,12,'Prove business readiness and user adoption','SIT report, UAT sign-off, training pack','Build + migration ready','UAT signed off'],
        ['9. Cutover & Go-Live',12,13,'Move safely into production','Cutover plan, rollback plan, go-live report','Go-live readiness approval','Go-live done'],
        ['10. Hypercare & Handover',13,16,'Stabilize and transition to BAU','Hypercare log, stabilization report, handover pack','Go-live complete','Hypercare closed'],
    ]


def timeline_rows(weeks=16):
    headers=['Workstream']+[f'W{i}' for i in range(1,weeks+1)]
    phases=[
        ('Initiation',1,1),('Requirements & Fit-Gap',2,4),('Solution Design',4,5),('Environment Setup',5,6),('Configuration',6,8),('Custom Build',6,10),('Data Migration',5,10),('Testing & Training',10,12),('Cutover & Go-Live',12,13),('Hypercare & Handover',13,16)
    ]
    rows=[headers]
    for name,start,end in phases:
        rows.append([name]+['■' if start<=w<=end else '' for w in range(1,weeks+1)])
    return rows


def critical_path_rows():
    return [
        ['CP No.','Critical Path Item','Why Critical','Predecessor','Success Measure'],
        [1,'Requirements sign-off','Prevents uncontrolled scope and design ambiguity','Kickoff + workshops complete','Signed scope baseline'],
        [2,'Solution blueprint approval','Configuration and build cannot proceed cleanly without it','Fit-gap complete','Approved blueprint'],
        [3,'Master data standardization','Migration and transactional integrity depend on clean masters','Master data design','Validated coding and structure'],
        [4,'Custom build completion','SIT/UAT cannot validate incomplete must-have scope','Blueprint + build backlog frozen','UAT-ready build'],
        [5,'Migration dry run validation','Go-live confidence depends on data rehearsal','Clean data + templates','Balanced validation report'],
        [6,'UAT sign-off','Business acceptance gate before production','SIT passed','Named user sign-off'],
        [7,'Cutover readiness approval','Production launch should not rely on improvisation','UAT passed + rollback ready','Approved checklist'],
    ]


def raid_rows():
    return [
        ['Type','ID','Description','Impact','Mitigation / Response','Owner','Due / Review','Status'],
        ['Risk','R-001','Scope creep after build starts','High','Freeze must-have scope and route extras through CR log','Implementer PM','Weekly','Open'],
        ['Risk','R-002','Dirty or incomplete master data','High','Assign data owners early and run dry runs before UAT','Customer PM','Weekly','Open'],
        ['Risk','R-003','Slow business decisions on workflows/codes','High','Maintain decision log and sponsor escalation path','Sponsor','Weekly','Open'],
        ['Risk','R-004','Weak UAT participation','High','Nominate users by module with signed test schedule','Customer PM','Weekly','Open'],
        ['Assumption','A-001','Customer will provide named key users','High','Confirm during kickoff','Customer PM','Week 1','Open'],
        ['Issue','I-001','Legacy data quality unknown until profiling','Medium','Profile sample data by week 3','Data Owner','Week 3','Open'],
        ['Dependency','D-001','Infrastructure/domain/email readiness','Medium','Track with technical setup checklist','Technical Lead','Week 6','Open'],
    ]


def decision_log_rows():
    return [
        ['Decision ID','Decision Topic','Decision Required By','Decision Owner','Options / Notes','Status'],
        ['DEC-001','Phase 1 module scope',2,'Sponsor + Customer PM','Freeze must-haves vs phase 2','Open'],
        ['DEC-002','Item code and naming standards',4,'Business Owner + Consultant','Affects migration and reporting','Open'],
        ['DEC-003','Approval workflow thresholds',4,'Sponsor + Customer PM','Needed for workflow design','Open'],
        ['DEC-004','Stock valuation / accounting approach',5,'Finance Owner','Needed before finance config','Open'],
        ['DEC-005','Go-live cutover window',12,'Sponsor + PMs','Need business freeze timing','Open'],
    ]


def module_plan_rows():
    return [
        ['Module','Primary Goals','Key Deliverables','Main Risks','Primary Owner'],
        ['Sales','Quotation-to-invoice control, approval visibility','Sales workflow, forms, reports, user training','Pricing/approval ambiguity','Functional Consultant'],
        ['Purchase','Controlled buying and receipt process','Purchase workflow, approval matrix, supplier data','Unclear approval authority','Functional Consultant'],
        ['Inventory','Warehouse accuracy and transaction discipline','Warehouses, stock movements, opening stock, reconciliation','Dirty item data, poor warehouse design','Inventory Lead'],
        ['Accounting','Financial posting accuracy and month-end control','COA, taxes, AR/AP opening, posting validation','COA confusion, opening balance mismatch','Finance Owner'],
        ['Manufacturing','Basic production planning and execution','BOMs, routings, work orders, FG movement','Incomplete BOM/routing data','Manufacturing Owner'],
    ]


def detailed_wbs_rows():
    rows=[['WBS Code','Phase','Task','Detailed Description','Owner','Predecessor','Start Week','End Week','Duration (Weeks)','Deliverable','Acceptance Criteria','Status']]
    data=[
        ('1.0','Initiation','Create project charter','Document objectives, scope, exclusions, governance, and success criteria','Implementer PM','-',1,1,1,'Project Charter','Sponsor and PMs approve charter','Planned'),
        ('1.1','Initiation','Assign project owners','Name sponsor, customer PM, implementer PM, module key users, data owner','Customer PM','1.0',1,1,1,'Stakeholder Register','Named owners recorded for all core roles','Planned'),
        ('1.2','Initiation','Set governance cadence','Define weekly PM meeting, steering cadence, escalation path, reporting','Implementer PM','1.0',1,1,1,'Governance Plan','Meeting cadence accepted by both sides','Planned'),
        ('1.3','Initiation','Open RAID and action logs','Create issue/risk/dependency/action registers','Implementer PM','1.0',1,1,1,'RAID Log','Logs available and in use','Planned'),
        ('2.0','Requirements','Run sales workshop','Capture lead-to-order and billing flow, approvals, outputs','Functional Consultant','1.2',2,2,1,'Sales Workshop Notes','Process and pain points documented','Planned'),
        ('2.1','Requirements','Run purchase workshop','Capture requisition, PO, receipt, invoice, approvals','Functional Consultant','1.2',2,2,1,'Purchase Workshop Notes','Process and exception cases documented','Planned'),
        ('2.2','Requirements','Run inventory workshop','Capture warehouse structure, transactions, reconciliation approach','Functional Consultant','1.2',2,3,2,'Inventory Workshop Notes','Warehouse and movement rules documented','Planned'),
        ('2.3','Requirements','Run accounting workshop','Capture taxes, COA, posting rules, reports, close controls','Functional Consultant','1.2',3,3,1,'Accounting Workshop Notes','Finance rules documented','Planned'),
        ('2.4','Requirements','Run manufacturing workshop','Capture BOM, routing, production issue/receipt, planning assumptions','Functional Consultant','1.2',3,3,1,'Manufacturing Workshop Notes','Manufacturing scope documented','Planned'),
        ('2.5','Requirements','Map current-state and to-be process','Turn workshops into process maps and future-state design','Solution Architect','2.0,2.1,2.2,2.3,2.4',3,4,2,'Process Maps','To-be process approved by module owners','Planned'),
        ('2.6','Requirements','Perform fit-gap analysis','Classify standard/config/custom/phase-2 requirements','Solution Architect','2.5',3,4,2,'Fit-Gap Matrix','Every major requirement classified','Planned'),
        ('2.7','Requirements','Define master data standards','Lock item code, UOM, warehouses, customer/supplier structure, COA','Customer Data Owner + Consultant','2.5',3,4,2,'Master Data Design','Standards approved for migration use','Planned'),
        ('2.8','Requirements','Sign off requirements','Approve scope baseline for phase 1','Customer PM + Key Users','2.6,2.7',4,4,1,'Signed Requirement Pack','Signed-off must-have scope','Planned'),
        ('3.0','Solution Design','Prepare solution blueprint','Document modules, workflows, roles, reports, integrations, migration strategy','Solution Architect','2.8',4,5,2,'Solution Blueprint','Blueprint approved by sponsor/PMs','Planned'),
        ('3.1','Solution Design','Define permission matrix','Align roles, approvers, and segregation of duties','Functional Lead','2.8',4,5,2,'Role Matrix','Role definitions accepted by business owners','Planned'),
        ('3.2','Solution Design','Prepare test strategy','Define SIT/UAT entry-exit criteria and scenario ownership','Implementer PM','3.0',5,5,1,'Test Strategy','Named test owners and gates defined','Planned'),
        ('3.3','Solution Design','Prepare cutover strategy','Define migration sequence, validation, rollback, hypercare model','Implementer PM','3.0',5,5,1,'Cutover Strategy','Go-live approach approved in principle','Planned'),
        ('4.0','Environment','Provision DEV/UAT/PROD','Prepare servers/sites and baseline access','Technical Lead','3.0',5,6,2,'Ready Environments','DEV/UAT/PROD accessible to project team','Planned'),
        ('4.1','Environment','Configure backups and SSL','Set technical controls for recovery and access','Technical Lead','4.0',5,6,2,'Technical Setup Checklist','Backup/SSL tested or verified','Planned'),
        ('4.2','Environment','Configure email and notifications baseline','Enable delivery of workflow alerts and basic messaging','Technical Lead','4.0',6,6,1,'Email Baseline','Test notifications sent successfully','Planned'),
        ('5.0','Configuration','Configure finance foundation','Set company, fiscal year, taxes, COA, cost centers','Functional Consultant','3.0',6,7,2,'Finance Base Config','Finance owner validates design intent','Planned'),
        ('5.1','Configuration','Configure sales process','Set quotation, sales order, delivery, invoicing basics','Functional Consultant','3.0',6,8,3,'Sales Config','Core sales flow demo passes','Planned'),
        ('5.2','Configuration','Configure purchase process','Set PR/PO/receipt/invoice basics and approvals','Functional Consultant','3.0',6,8,3,'Purchase Config','Core purchase flow demo passes','Planned'),
        ('5.3','Configuration','Configure inventory process','Set warehouses, movement rules, stock settings','Functional Consultant','3.0',6,8,3,'Inventory Config','Inventory test transactions pass','Planned'),
        ('5.4','Configuration','Configure manufacturing basics','Set BOM/routing/work order assumptions','Functional Consultant','3.0',7,8,2,'Manufacturing Config','Basic production cycle demo passes','Planned'),
        ('5.5','Configuration','Configure print and notification outputs','Prepare user-facing print formats and alerts','Functional Consultant','5.1,5.2,5.3',7,8,2,'Output Templates','Key forms available for UAT','Planned'),
        ('6.0','Custom Build','Develop must-have custom fields/scripts','Build approved UI and validation changes','Technical Lead','3.0',6,9,4,'Custom Build Part 1','Approved backlog items deployed to UAT','Planned'),
        ('6.1','Custom Build','Develop reports and dashboards','Build high-priority operational and management outputs','Technical Lead','3.0',7,10,4,'Reporting Pack','Critical reports available for UAT','Planned'),
        ('6.2','Custom Build','Develop integrations/automations','Implement approved external/API automations','Technical Lead','3.0',7,10,4,'Integration Pack','Integration test evidence available','Planned'),
        ('6.3','Custom Build','Internal build QA','Run developer/consultant functional checks before SIT','Technical Team','6.0,6.1,6.2',9,10,2,'Internal QA Log','No critical unresolved defects entering SIT','Planned'),
        ('7.0','Data Migration','Prepare migration templates','Issue templates for master/opening data','Data Consultant','2.7',5,6,2,'Migration Templates','Templates accepted by data owners','Planned'),
        ('7.1','Data Migration','Profile legacy data','Assess completeness, duplicates, code quality, mapping issues','Data Owner','7.0',6,7,2,'Data Profiling Report','Top data issues identified and assigned','Planned'),
        ('7.2','Data Migration','Cleanse master data','Normalize items, customers, suppliers, UOM, warehouses','Customer Data Owner','7.1',6,9,4,'Clean Master Data','Core masters pass validation rules','Planned'),
        ('7.3','Data Migration','Prepare opening balances and open docs','Compile stock, AR/AP, bank, open SO/PO/WO as needed','Accounting Owner + Data Consultant','7.1',8,10,3,'Opening Data Pack','Totals reconcile to source records','Planned'),
        ('7.4','Data Migration','Execute migration dry run','Load data to UAT and validate transaction readiness','Data Consultant + Key Users','7.2,7.3',9,10,2,'Dry Run Validation','Errors logged; balances validated','Planned'),
        ('8.0','Testing','Run SIT scenarios','Test end-to-end module and cross-module flows','Implementer Team','5.5,6.3,7.4',10,11,2,'SIT Report','Critical scenarios pass or defects assigned','Planned'),
        ('8.1','Testing','Fix SIT defects','Resolve blocking issues before business UAT','Technical Lead','8.0',10,11,2,'Defect Resolution Log','No critical blockers remain for UAT','Planned'),
        ('8.2','Testing','Run UAT by named users','Business validates real scenarios and exceptions','Customer PM + Key Users','8.1',11,12,2,'UAT Sign-off','Named user sign-off captured','Planned'),
        ('8.3','Testing','Train key users','Deliver process-led training to module owners','Functional Consultant','5.5',11,12,2,'Key User Training Pack','Key users can execute core flows','Planned'),
        ('8.4','Testing','Train end users','Deliver role-based end user training and quick guides','Functional Consultant','8.3',12,12,1,'End User Readiness','Attendance and guides completed','Planned'),
        ('9.0','Cutover','Prepare detailed cutover checklist','Break down final migration, access, validation, and communications','Implementer PM','8.2',12,13,2,'Cutover Checklist','Every cutover step has owner and timing','Planned'),
        ('9.1','Cutover','Prepare rollback and support plan','Define fallback path and war-room support model','Implementer PM','9.0',12,13,2,'Rollback + Hypercare Plan','Rollback path reviewed and accepted','Planned'),
        ('9.2','Cutover','Approve go-live readiness','Formal gate review across PM, sponsor, and leads','Sponsor + PMs','9.0,9.1',13,13,1,'Go-Live Approval','All critical gates approved','Planned'),
        ('9.3','Cutover','Execute production migration and launch','Load final data, validate controls, run first live transactions','Project Team','9.2',13,13,1,'Go-Live Report','Production launch complete','Planned'),
        ('10.0','Hypercare','Run hypercare triage','Support live users and manage P1/P2 defects','Support Lead + Key Users','9.3',13,14,2,'Hypercare Log','Critical live issues controlled','Planned'),
        ('10.1','Hypercare','Tune workflow/report/permission gaps','Resolve urgent operational gaps discovered in live usage','Support Lead','10.0',14,16,3,'Stabilization Report','Core operations stable','Planned'),
        ('10.2','Hypercare','Handover to BAU and phase 2 backlog','Transition support model and capture deferred enhancements','Implementer PM + Customer Owner','10.1',16,16,1,'Handover Pack','BAU owner accepts handover','Planned'),
    ]
    for item in data:
        rows.append(list(item))
    return rows


def dependency_rows():
    return [
        ['Dependency ID','Task / Deliverable','Depends On','Dependency Type','Impact if Delayed','Owner'],
        ['DEP-001','Requirements sign-off','Workshops + fit-gap completion','Finish-to-start','Blueprint/config/build delay','Customer PM'],
        ['DEP-002','Core configuration','Blueprint approval','Finish-to-start','Testing schedule slips','Functional Lead'],
        ['DEP-003','Migration dry run','Master data standards + clean data','Finish-to-start','UAT and go-live risk increase','Data Owner'],
        ['DEP-004','UAT','SIT pass + UAT-ready build','Finish-to-start','Go-live gate blocked','Customer PM'],
        ['DEP-005','Go-live','UAT sign-off + cutover approval','Finish-to-start','Production launch blocked','Sponsor + PMs'],
    ]


def deliverables_rows():
    return [
        ['Deliverable ID','Deliverable','Owner','Target Week','Acceptance Owner','Acceptance Standard'],
        ['DEL-001','Project Charter','Implementer PM',1,'Sponsor','Signed charter and governance'],
        ['DEL-002','Requirement & Fit-Gap Pack','Solution Architect',4,'Customer PM','Approved scope baseline'],
        ['DEL-003','Solution Blueprint','Solution Architect',5,'Sponsor','Blueprint approved'],
        ['DEL-004','Configured Core System','Functional Lead',8,'Customer PM','Demo-ready core flows'],
        ['DEL-005','Custom Build Pack','Technical Lead',10,'Implementer PM','UAT-ready approved build'],
        ['DEL-006','Migration Validation Report','Data Consultant',10,'Data Owner','Reconciled validation evidence'],
        ['DEL-007','UAT Sign-off','Customer PM',12,'Sponsor','Signed business acceptance'],
        ['DEL-008','Cutover & Go-Live Pack','Implementer PM',13,'Sponsor','Go-live approval + launch report'],
        ['DEL-009','Stabilization & Handover Pack','Support Lead',16,'Customer Owner','Accepted handover'],
    ]


def uat_rows():
    return [
        ['UAT Scenario ID','Business Area','Scenario','Owner','Preconditions','Expected Result','Status'],
        ['UAT-001','Sales','Quotation to Sales Invoice','Sales Key User','Customer/item/price ready','Complete flow posts correctly','Planned'],
        ['UAT-002','Purchase','PO to Purchase Invoice','Purchase Key User','Supplier/item/tax ready','Complete flow posts correctly','Planned'],
        ['UAT-003','Inventory','Receipt/Issue/Transfer/Reconciliation','Inventory Key User','Warehouse and stock ready','Stock balances behave correctly','Planned'],
        ['UAT-004','Accounting','AR/AP posting and reports','Finance Key User','Opening balances loaded','Financial outputs reconcile','Planned'],
        ['UAT-005','Manufacturing','WO issue/receipt finished goods','Manufacturing Key User','BOM/routing ready','Production movement works','Planned'],
        ['UAT-006','Approvals','Multi-role workflow approvals','Customer PM','Roles configured','Approvers can act per matrix','Planned'],
    ]


def cutover_rows():
    return [
        ['Step No.','Cutover Step','Owner','When','Validation','Rollback Trigger','Status'],
        [1,'Freeze source data and confirm cutover window','Customer PM','Pre go-live','Freeze announcement confirmed','Source still changing','Planned'],
        [2,'Take backup/snapshot of production baseline','Technical Lead','Pre go-live','Backup verified','Backup failure','Planned'],
        [3,'Load final master and opening data','Data Consultant','Go-live day','Import logs clean','Critical import failure','Planned'],
        [4,'Validate balances and critical counts','Finance Owner + Data Owner','Go-live day','Reconciliation complete','Balance mismatch unresolved','Planned'],
        [5,'Enable production user access','Technical Lead','Go-live day','Sample users log in successfully','Access/security issue','Planned'],
        [6,'Execute first live transactions','Key Users','Go-live day','Core transactions complete','P1 blocking issue','Planned'],
        [7,'Open hypercare war-room support','Support Lead','Go-live day','Support roster active','Support unavailable','Planned'],
    ]


def hypercare_rows():
    return [
        ['Area','Objective','Owner','Target Window','Success Measure'],
        ['Support triage','Resolve P1/P2 live issues quickly','Support Lead','Week 13-14','Critical issues responded within agreed SLA'],
        ['User adoption','Ensure key users run real work in system','Customer PM','Week 13-15','All target teams active in ERPNext'],
        ['Reporting tune-up','Fix urgent reporting/print gaps','Functional Lead','Week 14-16','Critical business outputs accepted'],
        ['Permission tuning','Correct access gaps without breaking control','Technical/Functional Leads','Week 13-15','No blocking access defects'],
        ['Phase 2 backlog','Capture deferred enhancements systematically','Implementer PM','Week 15-16','Prioritized backlog agreed'],
    ]


def action_tracker_rows():
    return [
        ['Action ID','Action','Owner','Due Week','Priority','Status','Notes'],
        ['ACT-001','Assign named key users and data owner','Customer PM',1,'High','Open','Critical kickoff prerequisite'],
        ['ACT-002','Freeze phase 1 must-have scope','Sponsor + PMs',4,'High','Open','Control scope creep'],
        ['ACT-003','Approve item/master data standards','Business Owner',4,'High','Open','Blocks migration template finalization'],
        ['ACT-004','Validate dry-run balances and exceptions','Data Owner',10,'High','Open','Blocks UAT confidence'],
        ['ACT-005','Approve cutover and rollback plan','Sponsor + PMs',13,'High','Open','Mandatory go-live gate'],
    ]


def build_workbook(project_name, customer_name, output_path):
    wb=new_workbook()
    add_sheet_from_rows(wb,'Executive Summary',executive_summary_rows(project_name, customer_name, 16),header_fill=GREEN,widths={1:24,2:100})
    add_sheet_from_rows(wb,'Scope & Assumptions',scope_rows(),header_fill=GREEN,widths={1:20,2:50,3:50})
    add_sheet_from_rows(wb,'Phase Plan',phase_plan_rows(),header_fill=BLUE,widths={1:20,2:12,3:12,4:32,5:34,6:28,7:26})
    add_sheet_from_rows(wb,'Milestones',milestone_rows(),header_fill=BLUE,widths={1:12,2:32,3:10,4:24,5:40,6:12})
    add_sheet_from_rows(wb,'High-Level Timeline',timeline_rows(16),header_fill=PURPLE,widths={1:24, **{i:5 for i in range(2,18)}})
    add_sheet_from_rows(wb,'Critical Path',critical_path_rows(),header_fill=ORANGE,widths={1:10,2:28,3:38,4:26,5:24})
    add_sheet_from_rows(wb,'Module Plan',module_plan_rows(),header_fill=GREEN,widths={1:18,2:28,3:34,4:28,5:22})
    add_sheet_from_rows(wb,'RAID Log',raid_rows(),header_fill=ORANGE,widths={1:12,2:10,3:34,4:12,5:36,6:18,7:14,8:12})
    add_sheet_from_rows(wb,'Decision Log',decision_log_rows(),header_fill=ORANGE,widths={1:12,2:30,3:16,4:18,5:34,6:12})
    add_sheet_from_rows(wb,'Action Tracker',action_tracker_rows(),header_fill=BLUE,widths={1:12,2:40,3:22,4:12,5:10,6:10,7:28})

    for ws in wb.worksheets:
        if ws.title == 'Executive Summary':
            ws.sheet_view.showGridLines = True
            for cell in ws['A']:
                cell.font = Font(bold=True)
    save_workbook(wb, output_path)


def build_wbs_workbook(output_path):
    wb=new_workbook()
    add_sheet_from_rows(wb,'Detailed WBS',detailed_wbs_rows(),header_fill=BLUE,widths={1:10,2:18,3:26,4:40,5:22,6:18,7:12,8:12,9:14,10:24,11:32,12:12})
    add_sheet_from_rows(wb,'Dependencies',dependency_rows(),header_fill=ORANGE,widths={1:12,2:28,3:28,4:18,5:28,6:18})
    add_sheet_from_rows(wb,'Deliverables Register',deliverables_rows(),header_fill=GREEN,widths={1:12,2:28,3:20,4:12,5:18,6:28})
    add_sheet_from_rows(wb,'UAT Plan',uat_rows(),header_fill=PURPLE,widths={1:14,2:18,3:30,4:20,5:24,6:24,7:12})
    add_sheet_from_rows(wb,'Cutover Plan',cutover_rows(),header_fill=ORANGE,widths={1:10,2:34,3:18,4:14,5:26,6:24,7:12})
    add_sheet_from_rows(wb,'Hypercare Plan',hypercare_rows(),header_fill=GREEN,widths={1:18,2:34,3:20,4:18,5:30})
    save_workbook(wb, output_path)


def main():
    p=argparse.ArgumentParser(description='Generate a real ERPNext PM implementation Excel pack')
    p.add_argument('--pack-output', required=True)
    p.add_argument('--wbs-output', required=True)
    p.add_argument('--project-name', default='ERPNext Implementation Plan')
    p.add_argument('--customer-name', default='Customer')
    args=p.parse_args()
    build_workbook(args.project_name, args.customer_name, args.pack_output)
    build_wbs_workbook(args.wbs_output)
    print(args.pack_output)
    print(args.wbs_output)

if __name__ == '__main__':
    main()
