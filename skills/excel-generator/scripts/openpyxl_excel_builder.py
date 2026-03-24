#!/usr/bin/env python3
import json
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THIN = Side(style='thin', color='D9D9D9')


def _apply_table_style(ws, header_fill='1F4E78', freeze='A2', widths=None, autofilter=True):
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_align = Alignment(vertical='top', wrap_text=True)
    border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.freeze_panes = freeze
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            if cell.row == 1:
                cell.fill = PatternFill('solid', fgColor=header_fill)
                cell.font = header_font
                cell.alignment = header_align
            else:
                cell.alignment = body_align
    if autofilter and ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    if widths:
        for idx, width in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = width
    else:
        for c in range(1, ws.max_column + 1):
            max_len = 10
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    max_len = min(max(max_len, len(str(v)) + 2), 60)
            ws.column_dimensions[get_column_letter(c)].width = max_len


def add_sheet_from_rows(wb: Workbook, title: str, rows: List[List[Any]], header_fill='1F4E78', freeze='A2', widths=None, autofilter=True):
    ws = wb.create_sheet(title=title[:31])
    for row in rows:
        ws.append(row)
    _apply_table_style(ws, header_fill=header_fill, freeze=freeze, widths=widths, autofilter=autofilter)
    return ws


def new_workbook(remove_default=True):
    wb = Workbook()
    if remove_default and wb.active and wb.active.title == 'Sheet':
        wb.remove(wb.active)
    return wb


def save_workbook(wb: Workbook, output_path: str):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def load_json(path: str):
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)
